import time
import os
from openpyxl import Workbook, load_workbook

filename = "favorite_people.xlsx"

if os.path.exists(filename):
    wb = load_workbook(filename)
    ws = wb.active
else:
     wb = Workbook()
     ws = wb.active

     headers = ["ID", "First Name", "Last Name", "Birth Year", "Age"]
     ws.append(headers)
     wb.save(filename)

for i in range(1, 4, 1):
    os.system('cls')
    print(">>> Favorite Person Recorder <<<")
    print("Input three of your favorite person's information here!")
    print(f"{i-1} out of 3 person.")
    firstName = input("Input you favorite person or character's first name: ")

    lastName = input("Input their last name: ")

    while True:
        birthYear = input("Their birthyear: ")

        if birthYear.isdigit() == False:
            print("You must input a valid birth year")
            continue
        else:
            break

    def compute_age():
        age = 2026 - int(birthYear)

        return age

    def append_to_excel():
        new_id = ws.max_row
        data = [new_id, firstName, lastName, int(birthYear), compute_age()]

        ws.append(data)
        wb.save(filename)
        print("Your favorite person or character has been added successfully.")
        time.sleep(1)

    append_to_excel()
    os.system('cls')

print("All three persons have been saved.")

for row in ws.iter_rows(min_row=2, values_only="True"):
    print(" | ".join(map(str, row)))